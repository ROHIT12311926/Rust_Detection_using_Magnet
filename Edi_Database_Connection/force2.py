import serial
import openpyxl

from openpyxl import load_workbook
from datetime import datetime
import os

file_path = "production_data_testing.xlsx"  # Fixed name

# 🔹 Ask user details
material = input("Enter Material: ")
length = input("Enter Length: ")
width = input("Enter Width: ")

# 🔹 Check if file already exists
if os.path.exists(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    print("📂 Existing Excel loaded. Appending data...")
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MLX90393 Data"
    ws.append(["Time", "Material", "Length", "Width", "X", "Y", "Z", "Magnitude"])
    print("🆕 New Excel created with headers.")

# ⚙️ Connect to Arduino
arduino = serial.Serial('COM3', 115200, timeout=1)
print("✅ Connected to Arduino. Logging started...")

try:
    while True:
        line = arduino.readline().decode(errors='ignore').strip()

        if not line or line.startswith(("Calibrating", "Calibration")):
            continue

        parts = line.split(",")

        # Ignore header lines
        if parts[0] == "X" or line.startswith("X,"):
            continue

        # Expecting: X, Y, Z, Magnitude, (Label - ignored)
        if len(parts) == 5:  
            X, Y, Z, magnitude, _ = parts  # last value ignored

            try:
                X = float(X)
                Y = float(Y)
                Z = float(Z)
                magnitude = float(magnitude)

                # ⭐ Save only if 125 < magnitude < 250
                if 125 < magnitude < 250:

                    ws.append([
                        datetime.now().strftime("%H:%M:%S"),
                        material,
                        length,
                        width,
                        X,
                        Y,
                        Z,
                        magnitude
                    ])
                    wb.save(file_path)

                    print(f"💾 Saved -> X={X}, Y={Y}, Z={Z}, Mag={magnitude}")

                else:
                    print(f"⛔ Skipped (Magnitude out of range): {magnitude}")

            except ValueError:
                print("⚠️ Invalid numeric data skipped:", line)

        else:
            print("⚠️ Unexpected data format:", line)

except KeyboardInterrupt:
    
    print("\n🛑 Logging stopped. File saved.")
    wb.save(file_path)
    arduino.close()
