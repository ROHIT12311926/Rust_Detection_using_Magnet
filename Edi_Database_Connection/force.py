import serial
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os

file_path = "MLX90393_Data.xlsx"

# 🔹 Check if file already exists
if os.path.exists(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    print("📂 Existing Excel file loaded. Data will be appended.")
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MLX90393 Data"
    ws.append(["Time", "Material", "Dimension_Length", "Dimension_Width", "X", "Y", "Z", "Magnitude", "Label"])
    print("🆕 New Excel file created with headers.")

# ⚙️ Connect to Arduino
arduino = serial.Serial('COM4', 115200, timeout=1)  # ⚠️ Apna COM port check kar lena
print("✅ Connected to Arduino. Logging started...")

try:
    while True:
        line = arduino.readline().decode(errors='ignore').strip()

        # Ignore empty or calibration lines
        if not line or line.startswith(("Calibrating", "Calibration")):
            continue

        parts = line.split(",")

        # Ignore header lines from Arduino
        if parts[0] == "X" or line.startswith("X,"):
            continue

        # Expecting 5 values: X, Y, Z, Magnitude, Label
        if len(parts) == 5:
            X, Y, Z, magnitude, label = parts
            try:
                ws.append([
                    datetime.now().strftime("%H:%M:%S"),
                    "",  # Material -> manually enter later
                    "",  # Length
                    "",  # Width
                    float(X),
                    float(Y),
                    float(Z),
                    float(magnitude),
                    label
                ])
                wb.save(file_path)
                print(f"💾 Saved -> X={X}, Y={Y}, Z={Z}, Mag={magnitude}, Label={label}")

            except ValueError:
                print("⚠️ Invalid data skipped:", line)

        else:
            print("⚠️ Unexpected format:", line)

except KeyboardInterrupt:
    print("\n🛑 Logging stopped. File saved as MLX90393_Data.xlsx")
    wb.save(file_path)
    arduino.close()
